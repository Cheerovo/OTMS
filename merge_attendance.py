#!/usr/bin/env python3
"""
合并 cheer 考勤表 Excel 为一张明细 CSV

用法:
  python3 merge_attendance.py /path/to/excel/folder

输出:
  在 Excel 文件夹下生成 attendance_detail.csv

  包含：
    - 明细数据（所有部门 sheet 的打卡记录合并，去重去噪）
    - 月度汇总数据（汇总 sheet 的请假/加班/假期余额）

噪声过滤：
  自动跳过 "年假确认单""签字""调休确认单" 等非员工行

依赖:
  pip3 install openpyxl
"""

import sys
import os
import re
import csv
from glob import glob
from datetime import datetime, timedelta
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("请先安装 openpyxl: pip3 install openpyxl")
    sys.exit(1)

# ---------------------------------------------------------------------------
# 配置
# ---------------------------------------------------------------------------

# 需要跳过的 sheet 名称关键词（不区分大小写）
SKIP_SHEET_KW = ['汇总', '统计', 'summary', '目录', '说明', 'sheet', '透视表', '数据透视', 'pivot', '核对']

# 名字别名映射：繁体/英文 → 花名册标准名（统一匹配到同一个员工ID）
NAME_ALIASES = {
    '賴淑賢': '赖淑贤',
    'yvonne lai': '赖淑贤',
    'Yvonne Lai': '赖淑贤',
    'YVONNE LAI': '赖淑贤',
}

def normalize_name(name):
    """将别名统一为标准名"""
    n = name.strip()
    return NAME_ALIASES.get(n, n)

# 噪声姓名关键词（包含任一即跳过该行）
NOISE_NAME_KW = [
    '年假确认单', '调休确认单', '加班确认单', '签字', '确认人',
    '制表人', '审核人', '批准人', '合计', '总计', '小计',
    '说明', '备注：', '注：', '员工姓名',
]

# ---------------------------------------------------------------------------
# 工具函数
# ---------------------------------------------------------------------------

def is_noise(name):
    """判断该行是否为噪声（非员工数据行）"""
    if not name:
        return True
    s = str(name).strip()
    if not s:
        return True
    # 纯数字（可能是序号残留）
    if re.match(r'^\d+$', s):
        return True
    # 太短
    if len(s) <= 1:
        return True
    # 噪声关键词
    for kw in NOISE_NAME_KW:
        if kw in s:
            return True
    return False


def parse_detail_date(val):
    """
    解析明细 sheet 的日期列。
    支持格式：
      - "25-12-26 星期五"  -> "2025-12-26"
      - "2026-01-05"        -> "2026-01-05"
      - datetime 对象
      - Excel 序列号
    """
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    if isinstance(val, (int, float)):
        # Excel 序列号
        base = datetime(1899, 12, 30)
        try:
            return (base + timedelta(days=int(val))).strftime('%Y-%m-%d')
        except Exception:
            return None

    s = str(val).strip()
    if not s:
        return None

    # "25-12-26 星期五" 这种格式
    m = re.match(r'(\d{2})-(\d{2})-(\d{2})', s)
    if m:
        yy = int(m.group(1))
        mm = int(m.group(2))
        dd = int(m.group(3))
        # 假设 20xx 年
        full_year = 2000 + yy if yy < 70 else 1900 + yy
        try:
            return datetime(full_year, mm, dd).strftime('%Y-%m-%d')
        except ValueError:
            return None

    # 尝试标准格式
    for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%m/%d/%Y', '%Y.%m.%d',
                '%Y年%m月%d日', '%d/%m/%Y']:
        try:
            return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue

    return None


def fmt_time(val):
    """格式化时间为 HH:MM"""
    if val is None:
        return ''
    if isinstance(val, datetime):
        return val.strftime('%H:%M')
    s = str(val).strip()
    # 已经是 "05:41" 格式
    if re.match(r'^\d{1,2}:\d{2}$', s):
        return s
    return s


def fmt_str(val):
    """安全转字符串"""
    if val is None:
        return ''
    s = str(val).strip()
    return s


# ---------------------------------------------------------------------------
# 处理单个 Excel 文件
# ---------------------------------------------------------------------------

def find_header_row(ws):
    """
    扫描前 5 行，找到包含"姓名"+"工作日"或"打卡"的标题行。
    返回 (header_row, headers_list) 或 (None, [])
    """
    max_col = ws.max_column or 1
    for r in range(1, min(6, ws.max_row + 1)):
        headers = []
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            headers.append(str(v).strip() if v else '')
        # 判断：必须同时有 姓名 + (工作日 或 打卡)
        has_name = any('姓名' in h for h in headers)
        has_date = any('工作日' in h for h in headers)
        has_clock = any('打卡' in h for h in headers)
        if has_name and (has_date or has_clock):
            return r, headers
    return None, []


def process_file(filepath):
    """
    返回 (detail_rows, summary_rows)
    """
    detail_rows = []
    summary_rows = []
    basename = os.path.basename(filepath)

    wb = openpyxl.load_workbook(filepath, data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row < 2:
            continue

        # 跳过辅助 sheet（透视表、核对等是 Excel 自动生成的冗余数据）
        sn_lower = sheet_name.lower()
        if any(kw in sn_lower for kw in ['透视表', '数据透视', 'pivot', '核对', '核对表']):
            continue

        # 查找表头行（可能是第 1/2/3 行）
        header_row, headers = find_header_row(ws)

        # 判断是否为汇总 sheet
        is_summary = any(kw in sn_lower for kw in ['汇总', 'summary'])
        # 也通过表头判断：含"序号""员工姓名"且不含"打卡"→ 汇总
        if header_row and not is_summary:
            header_str = ' '.join(headers)
            if '序号' in header_str and '员工姓名' in header_str and '打卡' not in header_str:
                is_summary = True

        if is_summary:
            rows = parse_summary_sheet(ws, basename, sheet_name)
            summary_rows.extend(rows)
        elif header_row:
            rows = parse_detail_sheet(ws, headers, header_row, basename)
            detail_rows.extend(rows)

    wb.close()
    return detail_rows, summary_rows


def parse_detail_sheet(ws, headers, header_row, basename):
    """解析打卡明细 sheet"""
    def find_col(*keywords):
        for idx, h in enumerate(headers):
            h_clean = h.replace('\n', '').replace(' ', '')
            for kw in keywords:
                if kw in h_clean:
                    return idx
        return None

    name_col = find_col('姓名')
    date_col = find_col('工作日', '日期', '考勤日期')
    shift_col = find_col('班次')
    ci_col = find_col('打卡1时间', '上班打卡', '签到')
    cir_col = find_col('打卡1结果', '上班结果')
    co_col = find_col('打卡2时间', '下班打卡', '签退')
    cor_col = find_col('打卡2结果', '下班结果')
    approval_col = find_col('审批单', '审批', '备注')

    if name_col is None:
        return []

    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        name = ws.cell(row=r, column=name_col + 1).value
        if is_noise(name):
            continue
        name_str = normalize_name(str(name).strip())

        date_val = ws.cell(row=r, column=date_col + 1).value if date_col is not None else None
        date_str = parse_detail_date(date_val)
        if not date_str:
            continue

        row = {
            '姓名': name_str,
            '日期': date_str,
            '班次': fmt_str(ws.cell(row=r, column=shift_col + 1).value) if shift_col is not None else '',
            '上班打卡': fmt_time(ws.cell(row=r, column=ci_col + 1).value) if ci_col is not None else '',
            '上班结果': fmt_str(ws.cell(row=r, column=cir_col + 1).value) if cir_col is not None else '',
            '下班打卡': fmt_time(ws.cell(row=r, column=co_col + 1).value) if co_col is not None else '',
            '下班结果': fmt_str(ws.cell(row=r, column=cor_col + 1).value) if cor_col is not None else '',
            '审批备注': fmt_str(ws.cell(row=r, column=approval_col + 1).value) if approval_col is not None else '',
            '来源文件': basename,
        }
        rows.append(row)

    return rows


def parse_summary_sheet(ws, basename, sheet_name):
    """解析月度汇总 sheet"""
    # 汇总 sheet 的结构：前几行是元数据，需要找真正的表头行
    # 表头特征：包含"序号""员工姓名"
    header_row = None
    for r in range(1, min(10, ws.max_row + 1)):
        row_vals = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            row_vals.append(str(v).strip() if v else '')
        row_str = ' '.join(row_vals)
        if '员工姓名' in row_str or ('姓名' in row_str and '序号' in row_str):
            header_row = r
            break

    if header_row is None:
        return []

    # 读取表头
    headers = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        headers.append(str(v).strip() if v else '')

    def find_col(*keywords):
        for idx, h in enumerate(headers):
            h_clean = h.replace('\n', '').replace(' ', '')
            for kw in keywords:
                if kw in h_clean:
                    return idx
        return None

    name_col = find_col('员工姓名', '姓名')
    dept_col = find_col('部门')
    late_col = find_col('迟到')
    miss_col = find_col('缺卡')
    abnormal_col = find_col('考勤异常')
    ot_col = find_col('加班')
    annual_col = find_col('年假')
    lieu_col = find_col('调休')
    sick_col = find_col('病假')
    personal_col = find_col('事假')
    remark_col = find_col('备注')

    # 尝试找日期（搜索前几行 + 从文件名/Sheet名提取）
    date_str = ''
    # 1) 先在 row 1~3 中搜日期（datetime / "2026-01" / "1月"）
    import_month_from_sheet = None
    for r in range(1, min(4, ws.max_row + 1)):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, datetime):
                date_str = v.strftime('%Y-%m-%d')
                break
            sv = str(v).strip() if v else ''
            m = re.search(r'(\d{4})-(\d{1,2})-(\d{1,2})', sv)
            if m:
                date_str = '{}-{:02d}-{:02d}'.format(int(m.group(1)), int(m.group(2)), int(m.group(3)))
                break
            # "2月" 格式 → 先记下来，年份稍后从文件名补
            m2 = re.match(r'^(\d{1,2})月$', sv)
            if m2:
                import_month_from_sheet = int(m2.group(1))
                break
        if date_str:
            break
    # 2) 如果只有月份没有年份，从 Sheet 名或文件名提取年份
    if not date_str and import_month_from_sheet:
        # 尝试从 sheet 名找年份
        year_from_name = None
        m3 = re.search(r'(\d{4})', sheet_name)
        if not m3:
            m3 = re.search(r'(\d{4})', basename)
        if m3:
            year_from_name = int(m3.group(1))
        else:
            year_from_name = 2026  # 兜底
        date_str = '{}-{:02d}-01'.format(year_from_name, import_month_from_sheet)
    # 3) 兜底: 从文件名提取 "202602" → "2026-02-01"
    if not date_str:
        m4 = re.search(r'(\d{4})(\d{2})', basename)
        if m4:
            date_str = '{}-{}-01'.format(m4.group(1), m4.group(2))

    if name_col is None:
        return []

    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        name = ws.cell(row=r, column=name_col + 1).value
        if is_noise(name):
            continue
        name_str = normalize_name(str(name).strip())

        row = {
            '姓名': name_str,
            '部门': fmt_str(ws.cell(row=r, column=dept_col + 1).value) if dept_col is not None else '',
            '统计月份': date_str,
            '迟到次数': fmt_str(ws.cell(row=r, column=late_col + 1).value) if late_col is not None else '',
            '缺卡次数': fmt_str(ws.cell(row=r, column=miss_col + 1).value) if miss_col is not None else '',
            '考勤异常天数': fmt_str(ws.cell(row=r, column=abnormal_col + 1).value) if abnormal_col is not None else '',
            '加班小时': fmt_str(ws.cell(row=r, column=ot_col + 1).value) if ot_col is not None else '',
            '年假天数': fmt_str(ws.cell(row=r, column=annual_col + 1).value) if annual_col is not None else '',
            '调休天数': fmt_str(ws.cell(row=r, column=lieu_col + 1).value) if lieu_col is not None else '',
            '病假天数': fmt_str(ws.cell(row=r, column=sick_col + 1).value) if sick_col is not None else '',
            '事假天数': fmt_str(ws.cell(row=r, column=personal_col + 1).value) if personal_col is not None else '',
            '备注': fmt_str(ws.cell(row=r, column=remark_col + 1).value) if remark_col is not None else '',
            '来源文件': basename,
        }
        rows.append(row)

    return rows


# ---------------------------------------------------------------------------
# 主流程
# ---------------------------------------------------------------------------

def main():
    if len(sys.argv) < 2:
        print("用法: python3 merge_attendance.py <存放Excel的文件夹>")
        print("示例: python3 merge_attendance.py ~/Downloads/考勤数据/")
        sys.exit(1)

    folder = os.path.abspath(sys.argv[1])
    if not os.path.isdir(folder):
        print(f"错误: 文件夹不存在 → {folder}")
        sys.exit(1)

    files = sorted(glob(os.path.join(folder, '*.xlsx')))
    # 排除临时文件（~$ 开头）
    files = [f for f in files if not os.path.basename(f).startswith('~$')]

    if not files:
        print(f"错误: 文件夹内没有 .xlsx 文件 → {folder}")
        sys.exit(1)

    print(f"找到 {len(files)} 个 Excel 文件\n")

    all_detail = []
    all_summary = []
    file_detail_counts = []

    for fp in files:
        fname = os.path.basename(fp)
        detail, summary = process_file(fp)
        all_detail.extend(detail)
        all_summary.extend(summary)
        file_detail_counts.append((fname, len(detail), len(summary)))
        print(f"  {fname}")
        print(f"    打卡明细 {len(detail):>5} 条  |  月度汇总 {len(summary):>3} 条")

    # ---------- 去重 ----------
    # 明细：同一个人同一天保留一条（不同文件可能重叠）
    seen_detail = {}
    for r in all_detail:
        key = (r['姓名'], r['日期'])
        if key not in seen_detail:
            seen_detail[key] = r
    detail_unique = sorted(seen_detail.values(), key=lambda r: (r['日期'], r['姓名']))

    # 汇总：同一个人同一月份保留一条
    seen_summary = {}
    for r in all_summary:
        key = (r['姓名'], r['统计月份'])
        if key not in seen_summary:
            seen_summary[key] = r
    summary_unique = sorted(seen_summary.values(), key=lambda r: (r['统计月份'], r['姓名']))

    # ---------- 输出 ----------
    detail_path = os.path.join(folder, '考勤明细_打卡记录.csv')
    detail_fields = ['姓名', '日期', '班次', '上班打卡', '上班结果', '下班打卡', '下班结果', '审批备注', '来源文件']

    with open(detail_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=detail_fields, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(detail_unique)

    summary_path = os.path.join(folder, '考勤明细_月度汇总.csv')
    summary_fields = ['姓名', '部门', '统计月份', '迟到次数', '缺卡次数', '考勤异常天数',
                      '加班小时', '年假天数', '调休天数', '病假天数', '事假天数', '备注', '来源文件']

    with open(summary_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=summary_fields, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(summary_unique)

    # ---------- 统计 ----------
    dates = sorted(set(r['日期'] for r in detail_unique))
    names = sorted(set(r['姓名'] for r in detail_unique))

    print(f"\n{'='*50}")
    print(f"完成！")
    print(f"  原始打卡记录: {len(all_detail)} 条")
    print(f"  去重后:        {len(detail_unique)} 条")
    print(f"  覆盖日期:      {dates[0]} ~ {dates[-1]} ({len(dates)} 天)")
    print(f"  涉及人数:      {len(names)} 人")
    print(f"  月度汇总:      {len(summary_unique)} 条")
    print(f"")
    print(f"  输出文件:")
    print(f"    打卡明细 → {detail_path}")
    print(f"    月度汇总 → {summary_path}")


if __name__ == '__main__':
    main()
